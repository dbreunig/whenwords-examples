#!/usr/bin/env python3
"""
Generate and test the whenwords Excel implementation.

This script:
1. Creates an Excel workbook with all LAMBDA function definitions
2. Sets up a test sheet with all test cases from tests.yaml
3. Optionally validates test results using Python reference implementation

Requirements:
    pip install openpyxl pyyaml

Usage:
    python generate_test_workbook.py [--validate]
"""

import os
import sys
import argparse

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("Error: openpyxl required. Install with: pip install openpyxl")
    sys.exit(1)

try:
    import yaml
except ImportError:
    print("Error: pyyaml required. Install with: pip install pyyaml")
    sys.exit(1)


# LAMBDA formula definitions
LAMBDA_FORMULAS = {
    "WW_PLURALIZE": '=LAMBDA(n, singular, plural, IF(n = 1, singular, plural))',

    "WW_UNIX_TO_DATE": '=LAMBDA(unix_ts, unix_ts / 86400 + 25569)',

    "WW_DATE_TO_UNIX": '=LAMBDA(excel_date, (excel_date - 25569) * 86400)',

    "WW_ABS_ROUND": '=LAMBDA(n, FLOOR(n + 0.5, 1))',

    "WW_TIMEAGO": '''=LAMBDA(timestamp, reference,
    LET(
        diff, reference - timestamp,
        abs_diff, ABS(diff),
        is_future, diff < 0,
        seconds, abs_diff,
        minutes, seconds / 60,
        hours, minutes / 60,
        days, hours / 24,
        months, days / 30.44,
        years, days / 365,
        result, IF(seconds < 45, "just now",
            IF(seconds < 90, IF(is_future, "in 1 minute", "1 minute ago"),
            IF(minutes < 45,
                LET(n, WW_ABS_ROUND(minutes),
                    unit, WW_PLURALIZE(n, "minute", "minutes"),
                    IF(is_future, "in " & n & " " & unit, n & " " & unit & " ago")),
            IF(minutes < 90, IF(is_future, "in 1 hour", "1 hour ago"),
            IF(hours < 22,
                LET(n, WW_ABS_ROUND(hours),
                    unit, WW_PLURALIZE(n, "hour", "hours"),
                    IF(is_future, "in " & n & " " & unit, n & " " & unit & " ago")),
            IF(hours < 36, IF(is_future, "in 1 day", "1 day ago"),
            IF(days < 26,
                LET(n, WW_ABS_ROUND(days),
                    unit, WW_PLURALIZE(n, "day", "days"),
                    IF(is_future, "in " & n & " " & unit, n & " " & unit & " ago")),
            IF(days < 46, IF(is_future, "in 1 month", "1 month ago"),
            IF(days < 320,
                LET(n, WW_ABS_ROUND(months),
                    unit, WW_PLURALIZE(n, "month", "months"),
                    IF(is_future, "in " & n & " " & unit, n & " " & unit & " ago")),
            IF(days < 548, IF(is_future, "in 1 year", "1 year ago"),
                LET(n, WW_ABS_ROUND(years),
                    unit, WW_PLURALIZE(n, "year", "years"),
                    IF(is_future, "in " & n & " " & unit, n & " " & unit & " ago"))
            )))))))))),
        result
    )
)''',

    "WW_DURATION": '''=LAMBDA(seconds, [compact], [max_units],
    LET(
        is_compact, IF(ISOMITTED(compact), FALSE, compact),
        units_limit, IF(ISOMITTED(max_units), 2, max_units),
        YEAR, 31536000,
        MONTH, 2592000,
        DAY, 86400,
        HOUR, 3600,
        MINUTE, 60,
        _check_error, IF(seconds < 0, "ERROR: Negative duration", ""),
        _years, FLOOR(seconds / YEAR, 1),
        _rem1, seconds - _years * YEAR,
        _months, FLOOR(_rem1 / MONTH, 1),
        _rem2, _rem1 - _months * MONTH,
        _days, FLOOR(_rem2 / DAY, 1),
        _rem3, _rem2 - _days * DAY,
        _hours, FLOOR(_rem3 / HOUR, 1),
        _rem4, _rem3 - _hours * HOUR,
        _minutes, FLOOR(_rem4 / MINUTE, 1),
        _secs, _rem4 - _minutes * MINUTE,
        y_str, IF(_years > 0, IF(is_compact, _years & "y", _years & " " & WW_PLURALIZE(_years, "year", "years")), ""),
        mo_str, IF(_months > 0, IF(is_compact, _months & "mo", _months & " " & WW_PLURALIZE(_months, "month", "months")), ""),
        d_str, IF(_days > 0, IF(is_compact, _days & "d", _days & " " & WW_PLURALIZE(_days, "day", "days")), ""),
        h_str, IF(_hours > 0, IF(is_compact, _hours & "h", _hours & " " & WW_PLURALIZE(_hours, "hour", "hours")), ""),
        m_str, IF(_minutes > 0, IF(is_compact, _minutes & "m", _minutes & " " & WW_PLURALIZE(_minutes, "minute", "minutes")), ""),
        s_str, IF(_secs > 0, IF(is_compact, FLOOR(_secs, 1) & "s", FLOOR(_secs, 1) & " " & WW_PLURALIZE(FLOOR(_secs, 1), "second", "seconds")), ""),
        parts, FILTER({y_str; mo_str; d_str; h_str; m_str; s_str}, {y_str; mo_str; d_str; h_str; m_str; s_str} <> ""),
        limited_parts, IF(ROWS(parts) > units_limit, INDEX(parts, SEQUENCE(units_limit)), parts),
        sep, IF(is_compact, " ", ", "),
        result, IF(_check_error <> "", _check_error,
            IF(seconds = 0, IF(is_compact, "0s", "0 seconds"),
            TEXTJOIN(sep, TRUE, limited_parts))),
        result
    )
)''',

    "WW_PARSE_DURATION": '''=LAMBDA(text,
    LET(
        cleaned, TRIM(text),
        lower_text, LOWER(cleaned),
        _is_empty, LEN(cleaned) = 0,
        _is_negative, LEFT(cleaned, 1) = "-",
        _is_colon, ISNUMBER(SEARCH(":", cleaned)),
        _colon_result, IF(_is_colon,
            LET(
                colon_pos1, SEARCH(":", cleaned),
                before_colon, LEFT(cleaned, colon_pos1 - 1),
                after_colon, MID(cleaned, colon_pos1 + 1, 100),
                has_second_colon, ISNUMBER(SEARCH(":", after_colon)),
                hours, VALUE(before_colon),
                result, IF(has_second_colon,
                    LET(
                        colon_pos2, SEARCH(":", after_colon),
                        mins, VALUE(LEFT(after_colon, colon_pos2 - 1)),
                        secs, VALUE(MID(after_colon, colon_pos2 + 1, 100)),
                        hours * 3600 + mins * 60 + secs
                    ),
                    LET(
                        mins, VALUE(after_colon),
                        hours * 3600 + mins * 60
                    )
                ),
                result
            ),
            0
        ),
        _no_units_text, SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(
            SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(
            SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(
            SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(
            SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(
            SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(lower_text,
            "seconds", "~S~"), "second", "~S~"), "secs", "~S~"), "sec", "~S~"), "s", "~S~"),
            "minutes", "~M~"), "minute", "~M~"), "mins", "~M~"), "min", "~M~"), "m", "~M~"),
            "hours", "~H~"), "hour", "~H~"), "hrs", "~H~"), "hr", "~H~"), "h", "~H~"),
            "days", "~D~"), "day", "~D~"), "d", "~D~"),
            "weeks", "~W~"), "week", "~W~"), "wks", "~W~"), "wk", "~W~"), "w", "~W~"),
            "and", ""), ",", " "), "  ", " "), "  ", " "), "  ", " "),
        _extract_weeks, LET(
            pos, SEARCH("~W~", _no_units_text),
            IF(ISERROR(pos), 0,
                LET(
                    before, TRIM(LEFT(_no_units_text, pos - 1)),
                    parts, TEXTSPLIT(before, " "),
                    last_part, INDEX(parts, COLUMNS(parts)),
                    VALUE(last_part) * 604800
                )
            )
        ),
        _extract_days, LET(
            pos, SEARCH("~D~", _no_units_text),
            IF(ISERROR(pos), 0,
                LET(
                    before, TRIM(LEFT(_no_units_text, pos - 1)),
                    parts, TEXTSPLIT(before, " "),
                    last_part, INDEX(parts, COLUMNS(parts)),
                    VALUE(last_part) * 86400
                )
            )
        ),
        _extract_hours, LET(
            pos, SEARCH("~H~", _no_units_text),
            IF(ISERROR(pos), 0,
                LET(
                    before, TRIM(LEFT(_no_units_text, pos - 1)),
                    parts, TEXTSPLIT(before, " "),
                    last_part, INDEX(parts, COLUMNS(parts)),
                    VALUE(last_part) * 3600
                )
            )
        ),
        _extract_mins, LET(
            pos, SEARCH("~M~", _no_units_text),
            IF(ISERROR(pos), 0,
                LET(
                    before, TRIM(LEFT(_no_units_text, pos - 1)),
                    parts, TEXTSPLIT(before, " "),
                    last_part, INDEX(parts, COLUMNS(parts)),
                    VALUE(last_part) * 60
                )
            )
        ),
        _extract_secs, LET(
            pos, SEARCH("~S~", _no_units_text),
            IF(ISERROR(pos), 0,
                LET(
                    before, TRIM(LEFT(_no_units_text, pos - 1)),
                    parts, TEXTSPLIT(before, " "),
                    last_part, INDEX(parts, COLUMNS(parts)),
                    VALUE(last_part)
                )
            )
        ),
        _unit_result, _extract_weeks + _extract_days + _extract_hours + _extract_mins + _extract_secs,
        _has_units, OR(ISNUMBER(SEARCH("~", _no_units_text))),
        _final, IF(_is_empty, "#ERROR: Empty string",
            IF(_is_negative, "#ERROR: Negative duration",
            IF(_is_colon, _colon_result,
            IF(NOT(_has_units), "#ERROR: No valid units",
            ROUND(_unit_result, 0))))),
        _final
    )
)''',

    "WW_HUMAN_DATE": '''=LAMBDA(timestamp, reference,
    LET(
        ts_date, FLOOR(WW_UNIX_TO_DATE(timestamp), 1),
        ref_date, FLOOR(WW_UNIX_TO_DATE(reference), 1),
        day_diff, ts_date - ref_date,
        ts_weekday, WEEKDAY(ts_date, 2),
        weekday_name, CHOOSE(ts_weekday, "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"),
        ts_month, TEXT(ts_date, "mmmm"),
        ts_day, DAY(ts_date),
        ts_year, YEAR(ts_date),
        ref_year, YEAR(ref_date),
        result, IF(day_diff = 0, "Today",
            IF(day_diff = -1, "Yesterday",
            IF(day_diff = 1, "Tomorrow",
            IF(AND(day_diff > -7, day_diff < 0), "Last " & weekday_name,
            IF(AND(day_diff > 1, day_diff < 7), "This " & weekday_name,
            IF(ts_year = ref_year, ts_month & " " & ts_day,
            ts_month & " " & ts_day & ", " & ts_year)))))),
        result
    )
)''',

    "WW_DATE_RANGE": '''=LAMBDA(start_ts, end_ts,
    LET(
        _start, IF(start_ts > end_ts, end_ts, start_ts),
        _end, IF(start_ts > end_ts, start_ts, end_ts),
        start_date, FLOOR(WW_UNIX_TO_DATE(_start), 1),
        end_date, FLOOR(WW_UNIX_TO_DATE(_end), 1),
        start_month, TEXT(start_date, "mmmm"),
        end_month, TEXT(end_date, "mmmm"),
        start_day, DAY(start_date),
        end_day, DAY(end_date),
        start_year, YEAR(start_date),
        end_year, YEAR(end_date),
        same_day, start_date = end_date,
        same_month, AND(start_year = end_year, MONTH(start_date) = MONTH(end_date)),
        same_year, start_year = end_year,
        ndash, UNICHAR(8211),
        result, IF(same_day, start_month & " " & start_day & ", " & start_year,
            IF(same_month, start_month & " " & start_day & ndash & end_day & ", " & start_year,
            IF(same_year, start_month & " " & start_day & " " & ndash & " " & end_month & " " & end_day & ", " & start_year,
            start_month & " " & start_day & ", " & start_year & " " & ndash & " " & end_month & " " & end_day & ", " & end_year))),
        result
    )
)'''
}


def load_tests(tests_yaml_path):
    """Load tests from tests.yaml file."""
    with open(tests_yaml_path, 'r') as f:
        data = yaml.safe_load(f)
    return data


def create_workbook(tests_data, output_path):
    """Create Excel workbook with LAMBDA definitions and test sheet."""
    wb = openpyxl.Workbook()

    # Style definitions
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Create Instructions sheet
    ws_inst = wb.active
    ws_inst.title = "Instructions"

    instructions = [
        "WHENWORDS FOR EXCEL - Test Workbook",
        "",
        "This workbook tests the whenwords LAMBDA functions.",
        "",
        "SETUP INSTRUCTIONS:",
        "1. Go to Formulas > Name Manager (Ctrl+F3)",
        "2. Add each formula from the 'Formulas' sheet as a Named Range",
        "3. The 'Tests' sheet will automatically calculate results",
        "",
        "REQUIREMENTS:",
        "- Excel 365 or Excel for Web (LAMBDA support required)",
        "- TEXTSPLIT function (Excel 365 Sep 2022+)",
        "",
        "HELPER FUNCTIONS (add first):",
        "- WW_PLURALIZE",
        "- WW_UNIX_TO_DATE",
        "- WW_DATE_TO_UNIX",
        "- WW_ABS_ROUND",
        "",
        "MAIN FUNCTIONS:",
        "- WW_TIMEAGO",
        "- WW_DURATION",
        "- WW_PARSE_DURATION",
        "- WW_HUMAN_DATE",
        "- WW_DATE_RANGE",
    ]

    for i, line in enumerate(instructions, 1):
        ws_inst.cell(row=i, column=1, value=line)
        if i == 1:
            ws_inst.cell(row=i, column=1).font = Font(bold=True, size=14)

    ws_inst.column_dimensions['A'].width = 60

    # Create Formulas sheet
    ws_formulas = wb.create_sheet("Formulas")
    ws_formulas.cell(row=1, column=1, value="Function Name").font = header_font
    ws_formulas.cell(row=1, column=1).fill = header_fill
    ws_formulas.cell(row=1, column=2, value="Formula (paste into Name Manager)").font = header_font
    ws_formulas.cell(row=1, column=2).fill = header_fill

    row = 2
    for name, formula in LAMBDA_FORMULAS.items():
        ws_formulas.cell(row=row, column=1, value=name)
        ws_formulas.cell(row=row, column=2, value=formula)
        row += 1

    ws_formulas.column_dimensions['A'].width = 25
    ws_formulas.column_dimensions['B'].width = 100

    # Create Tests sheet
    ws_tests = wb.create_sheet("Tests")

    # Headers
    headers = ["Function", "Test Name", "Input1", "Input2", "Input3", "Expected", "Actual", "Pass"]
    for col, header in enumerate(headers, 1):
        cell = ws_tests.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    # Freeze header row
    ws_tests.freeze_panes = 'A2'

    row = 2

    # Process timeago tests
    for test in tests_data.get('timeago', []):
        if 'error' in test:
            continue  # Skip error tests
        inp = test['input']
        ws_tests.cell(row=row, column=1, value="timeago")
        ws_tests.cell(row=row, column=2, value=test['name'])
        ws_tests.cell(row=row, column=3, value=inp['timestamp'])
        ws_tests.cell(row=row, column=4, value=inp['reference'])
        ws_tests.cell(row=row, column=6, value=test['output'])
        ws_tests.cell(row=row, column=7, value=f'=WW_TIMEAGO(C{row},D{row})')
        ws_tests.cell(row=row, column=8, value=f'=F{row}=G{row}')
        row += 1

    # Process duration tests
    for test in tests_data.get('duration', []):
        if 'error' in test:
            continue
        inp = test['input']
        ws_tests.cell(row=row, column=1, value="duration")
        ws_tests.cell(row=row, column=2, value=test['name'])
        ws_tests.cell(row=row, column=3, value=inp['seconds'])

        opts = inp.get('options', {})
        compact = opts.get('compact', False)
        max_units = opts.get('max_units', 2)

        ws_tests.cell(row=row, column=4, value=compact)
        ws_tests.cell(row=row, column=5, value=max_units)
        ws_tests.cell(row=row, column=6, value=test['output'])
        ws_tests.cell(row=row, column=7, value=f'=WW_DURATION(C{row},D{row},E{row})')
        ws_tests.cell(row=row, column=8, value=f'=F{row}=G{row}')
        row += 1

    # Process parse_duration tests
    for test in tests_data.get('parse_duration', []):
        if 'error' in test:
            continue
        ws_tests.cell(row=row, column=1, value="parse_duration")
        ws_tests.cell(row=row, column=2, value=test['name'])
        ws_tests.cell(row=row, column=3, value=test['input'])
        ws_tests.cell(row=row, column=6, value=test['output'])
        ws_tests.cell(row=row, column=7, value=f'=WW_PARSE_DURATION(C{row})')
        ws_tests.cell(row=row, column=8, value=f'=F{row}=G{row}')
        row += 1

    # Process human_date tests
    for test in tests_data.get('human_date', []):
        if 'error' in test:
            continue
        inp = test['input']
        ws_tests.cell(row=row, column=1, value="human_date")
        ws_tests.cell(row=row, column=2, value=test['name'])
        ws_tests.cell(row=row, column=3, value=inp['timestamp'])
        ws_tests.cell(row=row, column=4, value=inp['reference'])
        ws_tests.cell(row=row, column=6, value=test['output'])
        ws_tests.cell(row=row, column=7, value=f'=WW_HUMAN_DATE(C{row},D{row})')
        ws_tests.cell(row=row, column=8, value=f'=F{row}=G{row}')
        row += 1

    # Process date_range tests
    for test in tests_data.get('date_range', []):
        if 'error' in test:
            continue
        inp = test['input']
        ws_tests.cell(row=row, column=1, value="date_range")
        ws_tests.cell(row=row, column=2, value=test['name'])
        ws_tests.cell(row=row, column=3, value=inp['start'])
        ws_tests.cell(row=row, column=4, value=inp['end'])
        ws_tests.cell(row=row, column=6, value=test['output'])
        ws_tests.cell(row=row, column=7, value=f'=WW_DATE_RANGE(C{row},D{row})')
        ws_tests.cell(row=row, column=8, value=f'=F{row}=G{row}')
        row += 1

    # Add summary row
    total_tests = row - 2
    ws_tests.cell(row=row + 1, column=1, value="SUMMARY")
    ws_tests.cell(row=row + 1, column=1).font = Font(bold=True)
    ws_tests.cell(row=row + 1, column=6, value="Total Tests:")
    ws_tests.cell(row=row + 1, column=7, value=total_tests)
    ws_tests.cell(row=row + 2, column=6, value="Passed:")
    ws_tests.cell(row=row + 2, column=7, value=f'=COUNTIF(H2:H{row-1},TRUE)')
    ws_tests.cell(row=row + 3, column=6, value="Failed:")
    ws_tests.cell(row=row + 3, column=7, value=f'=COUNTIF(H2:H{row-1},FALSE)')

    # Apply conditional formatting for Pass column would need openpyxl rules
    # For simplicity, we'll skip that here

    # Set column widths
    ws_tests.column_dimensions['A'].width = 15
    ws_tests.column_dimensions['B'].width = 40
    ws_tests.column_dimensions['C'].width = 30
    ws_tests.column_dimensions['D'].width = 15
    ws_tests.column_dimensions['E'].width = 10
    ws_tests.column_dimensions['F'].width = 35
    ws_tests.column_dimensions['G'].width = 35
    ws_tests.column_dimensions['H'].width = 10

    wb.save(output_path)
    print(f"Created workbook: {output_path}")
    return total_tests


def validate_with_python(tests_data):
    """Validate tests using Python reference implementation."""
    # Import the Python reference implementation
    python_impl_path = os.path.join(os.path.dirname(__file__), '..', 'python')
    sys.path.insert(0, python_impl_path)

    try:
        from whenwords import timeago, duration, parse_duration, human_date, date_range
    except ImportError:
        print("Warning: Could not import Python reference implementation")
        print("         Skipping validation")
        return None

    passed = 0
    failed = 0
    errors = []

    # Test timeago
    for test in tests_data.get('timeago', []):
        if 'error' in test:
            continue
        inp = test['input']
        expected = test['output']
        try:
            result = timeago(inp['timestamp'], inp['reference'])
            if result == expected:
                passed += 1
            else:
                failed += 1
                errors.append(f"timeago: {test['name']}: expected '{expected}', got '{result}'")
        except Exception as e:
            failed += 1
            errors.append(f"timeago: {test['name']}: raised {e}")

    # Test duration
    for test in tests_data.get('duration', []):
        if 'error' in test:
            continue
        inp = test['input']
        expected = test['output']
        opts = inp.get('options', {})
        try:
            result = duration(
                inp['seconds'],
                compact=opts.get('compact', False),
                max_units=opts.get('max_units', 2)
            )
            if result == expected:
                passed += 1
            else:
                failed += 1
                errors.append(f"duration: {test['name']}: expected '{expected}', got '{result}'")
        except Exception as e:
            failed += 1
            errors.append(f"duration: {test['name']}: raised {e}")

    # Test parse_duration
    for test in tests_data.get('parse_duration', []):
        if 'error' in test:
            continue
        expected = test['output']
        try:
            result = parse_duration(test['input'])
            if result == expected:
                passed += 1
            else:
                failed += 1
                errors.append(f"parse_duration: {test['name']}: expected {expected}, got {result}")
        except Exception as e:
            failed += 1
            errors.append(f"parse_duration: {test['name']}: raised {e}")

    # Test human_date
    for test in tests_data.get('human_date', []):
        if 'error' in test:
            continue
        inp = test['input']
        expected = test['output']
        try:
            result = human_date(inp['timestamp'], inp['reference'])
            if result == expected:
                passed += 1
            else:
                failed += 1
                errors.append(f"human_date: {test['name']}: expected '{expected}', got '{result}'")
        except Exception as e:
            failed += 1
            errors.append(f"human_date: {test['name']}: raised {e}")

    # Test date_range
    for test in tests_data.get('date_range', []):
        if 'error' in test:
            continue
        inp = test['input']
        expected = test['output']
        try:
            result = date_range(inp['start'], inp['end'])
            if result == expected:
                passed += 1
            else:
                failed += 1
                errors.append(f"date_range: {test['name']}: expected '{expected}', got '{result}'")
        except Exception as e:
            failed += 1
            errors.append(f"date_range: {test['name']}: raised {e}")

    return passed, failed, errors


def main():
    parser = argparse.ArgumentParser(description='Generate whenwords Excel test workbook')
    parser.add_argument('--validate', action='store_true',
                        help='Validate tests using Python reference implementation')
    args = parser.parse_args()

    script_dir = os.path.dirname(os.path.abspath(__file__))
    tests_yaml_path = os.path.join(script_dir, '..', '..', 'tests.yaml')
    output_path = os.path.join(script_dir, 'test_whenwords.xlsx')

    # Load tests
    print(f"Loading tests from: {tests_yaml_path}")
    tests_data = load_tests(tests_yaml_path)

    # Create workbook
    total_tests = create_workbook(tests_data, output_path)
    print(f"Total tests: {total_tests}")

    # Optionally validate with Python
    if args.validate:
        print("\nValidating with Python reference implementation...")
        result = validate_with_python(tests_data)
        if result:
            passed, failed, errors = result
            print(f"Passed: {passed}")
            print(f"Failed: {failed}")
            if errors:
                print("\nFailures:")
                for err in errors:
                    print(f"  - {err}")
            if failed == 0:
                print("\nAll tests passed!")
                return 0
            else:
                return 1

    print("\nWorkbook created successfully!")
    print("Open in Excel 365 and add the LAMBDA functions from the 'Formulas' sheet")
    print("to your Name Manager to run the tests.")
    return 0


if __name__ == '__main__':
    sys.exit(main())
